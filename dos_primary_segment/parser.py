"""
DOS dataset parser. Stops at TRANSIT SUPERVISOR. Produces raw rows for packet building.
"""
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Optional, Tuple

# Sentinel: stop processing when a line contains this (exact phrase)
TRANSIT_SUPERVISOR = "TRANSIT SUPERVISOR"

# Primary condition keywords: if remainder contains these, set primary_condition_text (exclude packet)
# Excludes: pay-treatment notes like ".50 CTE E/B", "1.02 PAID AS CTE" — those describe how to code OT/LPI,
# not a condition that prevents the run. "e/b" and "paid as cte/ot" alone = pay instruction, not condition.
PRIMARY_CONDITION_KEYWORDS = {
    "sick continued", "sick not counted", "sick until", "open", "admin leave",
    "annual leave", "vacation", "fmla", "ttd", "workers comp", "training",
    "nda", "loa without pay", "miscellaneous", "shine", "not counted",
    "sick until (not counted)",
}


@dataclass
class RawRow:
    """One parsed DOS row (before inclusion/exclusion)."""
    paddle: str
    block: str
    shift_time_str: str
    actual_start_str: str
    actual_end_str: str
    scheduled_end_str: str
    emp_id: str
    employee_name: str
    alternate_driver_present: bool
    notes_text: str
    primary_condition_text: str
    source_line_index: int  # 0-based line number in DOS
    potential_bleed: bool = False  # True if notes contain another emp_id — possible PDF line merge
    alternate_emp_id: str = ""   # Set when alternate driver present
    alternate_name: str = ""      # Set when alternate driver present


# Pattern for "Name (emp_id)" — used for driver extraction and bleed detection
_DRIVER_ID_PATTERN = re.compile(r"\((\d+)\)")


def _extract_driver_ids_and_rest(line: str) -> Tuple[List[Tuple[str, str]], str]:
    """Find all 'Name (id)' patterns; return [(name, id), ...] and remainder."""
    rest = line
    drivers = []
    # Match "Something (digits)" - name can have spaces, digits is emp_id
    pattern = re.compile(r"([^(]+?)\((\d+)\)")
    for m in pattern.finditer(line):
        name = m.group(1).strip()
        eid = m.group(2).strip()
        drivers.append((name, eid))
    # Remainder: take everything after the last ")"
    last_close = line.rfind(")")
    if last_close >= 0:
        rest = line[last_close + 1 :].strip()
    else:
        rest = ""
    return drivers, rest


def _has_potential_bleed(rest: str, current_emp_id: str) -> bool:
    """
    Check if remainder contains another (emp_id) pattern — suggests PDF line merge/bleed.
    Notes should not contain another employee's ID; if it does, text may have jumped rows.
    """
    if not rest or not current_emp_id:
        return False
    for m in _DRIVER_ID_PATTERN.finditer(rest):
        other_id = m.group(1)
        if other_id != current_emp_id:
            return True
    return False


def _classify_remainder(rest: str) -> Tuple[str, str]:
    """Split remainder into notes_text and primary_condition_text. Conservative: any condition keyword -> primary."""
    rest_lower = rest.lower()
    for kw in PRIMARY_CONDITION_KEYWORDS:
        if kw in rest_lower:
            return rest, rest  # full rest as both; packet will be excluded for PRIMARY_CONDITION
    return rest, ""


def _parse_data_line(line: str, line_index: int) -> Optional[RawRow]:
    """
    Parse a single DOS data line. Returns None if line doesn't match expected format.
    Expected: paddle block shift_time hrs vehicle start end trim [primary_driver] [alternate] [notes/condition]
    """
    line = line.strip()
    if not line or line == TRANSIT_SUPERVISOR:
        return None

    shift_range_re = re.compile(r"(\d{1,2}:\d{2})-(\d{1,2}:\d{2})")
    shift_match = shift_range_re.search(line)
    if not shift_match:
        return None

    # Fixed columns: paddle block shift_time hrs [vehicle] start end trim
    # Some rows (e.g. 9003, EXB) omit vehicle: hrs then start end trim.
    fixed_re_with_vehicle = re.compile(
        r"^(\d+)\s+"           # paddle
        r"(\S+)\s+"            # block
        r"(\d{1,2}:\d{2}-\d{1,2}:\d{2})\s+"  # shift_time
        r"([\d.]+)\s+"         # hrs
        r"(\d+)\s+"            # vehicle
        r"(\d{1,2}:\d{2})\s+" # start
        r"(\d{1,2}:\d{2})\s+" # end
        r"([\d.]+)\s*"         # trim
    )
    fixed_re_no_vehicle = re.compile(
        r"^(\d+)\s+"           # paddle
        r"(\S+)\s+"            # block
        r"(\d{1,2}:\d{2}-\d{1,2}:\d{2})\s+"  # shift_time
        r"([\d.]+)\s+"         # hrs
        r"(\d{1,2}:\d{2})\s+" # start (no vehicle)
        r"(\d{1,2}:\d{2})\s+" # end
        r"([\d.]+)\s*"         # trim
    )
    m = fixed_re_with_vehicle.match(line)
    if m:
        paddle, block, shift_time_str, actual_start_str, actual_end_str = (
            m.group(1), m.group(2), m.group(3), m.group(6), m.group(7),
        )
    else:
        m = fixed_re_no_vehicle.match(line)
        if not m:
            return None
        paddle, block, shift_time_str, actual_start_str, actual_end_str = (
            m.group(1), m.group(2), m.group(3), m.group(5), m.group(6),
        )
    scheduled_end_str = shift_range_re.search(shift_time_str).group(2)

    rest = line[m.end() :].strip()
    drivers, rest = _extract_driver_ids_and_rest(rest)
    if not drivers:
        return None
    primary_name, emp_id = drivers[0]
    alternate_driver_present = len(drivers) >= 2
    alt_emp_id = drivers[1][1] if len(drivers) >= 2 else ""
    alt_name = drivers[1][0].strip() if len(drivers) >= 2 else ""
    notes_text, primary_condition_text = _classify_remainder(rest)
    potential_bleed = _has_potential_bleed(notes_text, emp_id)

    return RawRow(
        paddle=paddle,
        block=block,
        shift_time_str=shift_time_str,
        actual_start_str=actual_start_str,
        actual_end_str=actual_end_str,
        scheduled_end_str=scheduled_end_str,
        emp_id=emp_id,
        employee_name=primary_name.strip(),
        alternate_driver_present=alternate_driver_present,
        notes_text=notes_text,
        primary_condition_text=primary_condition_text,
        source_line_index=line_index,
        potential_bleed=potential_bleed,
        alternate_emp_id=alt_emp_id,
        alternate_name=alt_name,
    )


def extract_lines_from_pdf(pdf_path: Path) -> List[str]:
    """
    Extract text lines from PDF (one page).
    Uses pdfplumber extract_text() which returns lines in reading order.
    Assumption: one logical DOS row per extracted line — parser expects paddle, block, times,
    driver(s), and notes all on the same line. If the PDF layout merges or splits rows
    differently, words can "jump" between rows. We detect potential bleed when notes
    contain another employee's (id) and set potential_bleed for review.
    """
    try:
        import pdfplumber
    except ImportError:
        raise RuntimeError("pdfplumber is required for PDF input. pip install pdfplumber")
    with pdfplumber.open(pdf_path) as pdf:
        if not pdf.pages:
            return []
        text = pdf.pages[0].extract_text()
        if not text:
            return []
        return [ln.strip() for ln in text.splitlines()]


def extract_lines_from_text_file(path: Path) -> List[str]:
    """Read lines from a plain text file."""
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        return [ln.rstrip("\n\r") for ln in f.readlines()]


# Pattern: data rows start with paddle (4-5 digits) + block
_DATA_ROW_START = re.compile(r"^\d{4,5}\s+\S+")


def _is_note_continuation(line: str) -> bool:
    """
    True if line appears to be Notes column content for previous row (PDF layout).
    EXB notes like "1.30 PAID AS OT" or "SHINE - 0400-0514" appear on the following line.
    """
    line = line.strip()
    if not line or TRANSIT_SUPERVISOR in line:
        return False
    if _DATA_ROW_START.match(line):
        return False  # New data row, not a note
    return True


def parse_dos_lines(lines: List[str]) -> Tuple[List[RawRow], int, Optional[str]]:
    """
    Parse DOS lines. Stop at first line that contains TRANSIT SUPERVISOR.
    For PDF: Notes column can appear on following line(s); we merge those into notes_text.
    Returns: (list of RawRows, 1-based row index where we stopped, work_date from doc or None).
    """
    rows = []
    stopped_at_line_1based = 0
    work_date = None
    date_re = re.compile(r"(\d{1,2})/(\d{1,2})/(\d{2,4})")
    i = 0
    while i < len(lines):
        line = lines[i]
        if TRANSIT_SUPERVISOR in line:
            stopped_at_line_1based = i + 1
            break
        dm = date_re.search(line)
        if dm:
            mo, day, yr = dm.group(1), dm.group(2), dm.group(3)
            if len(yr) == 2:
                yr = "20" + yr
            work_date = f"{int(mo):02d}/{int(day):02d}/{yr}"
        raw = _parse_data_line(line, i)
        if raw is not None:
            # Merge following note continuation lines (PDF: Notes column on next line)
            j = i + 1
            extra_parts = []
            while j < len(lines) and _is_note_continuation(lines[j]):
                extra_parts.append(lines[j].strip())
                j += 1
            if extra_parts:
                raw.notes_text = (raw.notes_text + " " + " ".join(extra_parts)).strip()
            rows.append(raw)
            i = j  # Skip consumed continuation lines
        else:
            i += 1

    if stopped_at_line_1based == 0:
        stopped_at_line_1based = len(lines)

    return rows, stopped_at_line_1based, work_date


def _col(row: dict, *keys: str) -> str:
    """Get first matching column value (case-insensitive header match)."""
    for k in keys:
        for h, v in row.items():
            if h and k in str(h).lower() and v is not None:
                return str(v).strip()
    return ""


def _parse_driver_cell(cell: str) -> Tuple[str, str]:
    """Parse 'Name (emp_id)' -> (name, emp_id)."""
    m = re.search(r"([^(]*?)\s*\((\d+)\)\s*$", (cell or "").strip())
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return "", ""


# Trim spillover: when Trim bleeds into Primary Driver, we get "4.77 Adelaida Robledo (2964)"
_TRIM_SPILLOVER_RE = re.compile(r"^\d+(?:\.\d+)?\s+(.+)$")

def _clean_primary_driver_cell(trim_cell: str, primary_cell: str) -> str:
    cell = primary_cell.strip() if primary_cell else ""
    if not cell and trim_cell:
        cell = trim_cell.strip()
    if not cell:
        return ""
    m = _TRIM_SPILLOVER_RE.match(cell)
    return m.group(1).strip() if m else cell

def _strip_leading_trim(s: str) -> str:
    if not s:
        return s
    m = _TRIM_SPILLOVER_RE.match(s.strip())
    return m.group(1).strip() if m else s.strip()


def extract_raw_rows_from_csv(path: Path) -> Tuple[List[RawRow], int, Optional[str]]:
    """
    Load DOS from CSV with columns: Paddle, Block, Shift Time, Vehicle?, Start, End,
    Primary Driver, Alternate Driver, Notes, Primary Condition.
    """
    import csv
    rows_out = []
    work_date = None
    date_re = re.compile(r"(\d{1,2})/(\d{1,2})/(\d{2,4})")
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        reader = csv.DictReader(f)
        for i, row in enumerate(reader):
            # Stop at TRANSIT SUPERVISOR
            row_str = str(row).lower()
            if "transit" in row_str and "supervisor" in row_str:
                return rows_out, i + 1, work_date
            # Date from first row sometimes
            for v in row.values():
                if v and date_re.search(str(v)):
                    dm = date_re.search(str(v))
                    if dm:
                        mo, day, yr = dm.group(1), dm.group(2), dm.group(3)
                        if len(yr) == 2:
                            yr = "20" + yr
                        work_date = f"{int(mo):02d}/{int(day):02d}/{yr}"
                    break
            prim = _col(row, "primary driver", "primary_driver", "primary")
            if not prim:
                continue
            p_name, p_id = _parse_driver_cell(prim)
            if not p_id:
                continue
            shift = _col(row, "shift time", "shift_time", "shift")
            start = _col(row, "start")
            end = _col(row, "end")
            if not shift or not start or not end:
                continue
            # Scheduled end from shift range HH:MM-HH:MM
            shift_match = re.search(r"(\d{1,2}:\d{2})-(\d{1,2}:\d{2})", shift)
            sched_end = shift_match.group(2) if shift_match else end
            alt = _col(row, "alternate driver", "alternate_driver", "alternate")
            alt_name, alt_id = _parse_driver_cell(alt)
            notes = _col(row, "notes")
            cond = _col(row, "primary condition", "primary_condition", "condition")
            _, primary_cond = _classify_remainder(notes + " " + cond) if (notes or cond) else ("", "")
            raw = RawRow(
                paddle=_col(row, "paddle") or "",
                block=_col(row, "block") or "",
                shift_time_str=shift,
                actual_start_str=start,
                actual_end_str=end,
                scheduled_end_str=sched_end,
                emp_id=p_id,
                employee_name=p_name,
                alternate_driver_present=bool(alt_id),
                notes_text=notes,
                primary_condition_text=primary_cond,
                source_line_index=i,
                potential_bleed=_has_potential_bleed(notes, p_id),
                alternate_emp_id=alt_id,
                alternate_name=alt_name,
            )
            rows_out.append(raw)
    return rows_out, len(rows_out) + 1, work_date


def extract_raw_rows_from_xlsx(path: Path) -> Tuple[List[RawRow], int, Optional[str]]:
    """
    Load DOS from Excel. Expects columns: Paddle, Block, Shift Time, Start, End,
    Primary Driver, Alternate Driver, Notes, Primary Condition.
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl required for Excel. pip install openpyxl")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], 0, None
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    header_lower = [h.lower() for h in header]
    work_date = None
    date_re = re.compile(r"(\d{1,2})/(\d{1,2})/(\d{2,4})")
    rows_out = []
    for i, row in enumerate(rows[1:], start=1):
        if not row:
            continue
        row_dict = dict(zip(header, ["" if c is None else str(c).strip() for c in row]))
        row_str = " ".join(str(x) for x in row).lower()
        if "transit" in row_str and "supervisor" in row_str:
            return rows_out, i + 1, work_date
        def col(*keys):
            for k in keys:
                for j, h in enumerate(header_lower):
                    if k in h and j < len(row) and row[j] is not None:
                        return str(row[j]).strip()
            return ""
        prim = col("primary driver", "primary_driver", "primary")
        if not prim:
            continue
        p_name, p_id = _parse_driver_cell(prim)
        if not p_id:
            continue
        shift = col("shift time", "shift_time", "shift")
        start = col("start")
        end = col("end")
        if not shift or not start or not end:
            continue
        shift_match = re.search(r"(\d{1,2}:\d{2})-(\d{1,2}:\d{2})", shift)
        sched_end = shift_match.group(2) if shift_match else end
        alt = col("alternate driver", "alternate_driver", "alternate")
        alt_name, alt_id = _parse_driver_cell(alt)
        notes = col("notes")
        cond = col("primary condition", "primary_condition", "condition")
        _, primary_cond = _classify_remainder(notes + " " + cond) if (notes or cond) else ("", "")
        raw = RawRow(
            paddle=col("paddle"),
            block=col("block"),
            shift_time_str=shift,
            actual_start_str=start,
            actual_end_str=end,
            scheduled_end_str=sched_end,
            emp_id=p_id,
            employee_name=p_name,
            alternate_driver_present=bool(alt_id),
            notes_text=notes,
            primary_condition_text=primary_cond,
            source_line_index=i,
            potential_bleed=_has_potential_bleed(notes, p_id),
            alternate_emp_id=alt_id,
            alternate_name=alt_name,
        )
        rows_out.append(raw)
    return rows_out, len(rows_out) + 1, work_date


# Preliminary DOS: projected hours. Start and End columns ARE the shift (source of truth, nothing to compare to).
# PDF can have: paddle block start end hrs [rest] OR paddle block shift_time hrs [rest]
_PRELIM_RE = re.compile(
    r"^(\d+)\s+"           # paddle
    r"(\S+)\s+"            # block
    r"(\d{1,2}:\d{2}-\d{1,2}:\d{2})\s+"  # shift_time (range)
    r"([\d.]+)\s*"         # hrs
)
_PRELIM_START_END_RE = re.compile(  # paddle block ... start end ... (Start/End are the projected shift)
    r"^(\d+)\s+"           # paddle
    r"(\S+)\s+"            # block
    r"(?:(?:[\d.]+)\s+)*"  # optional hrs etc
    r"(\d{1,2}:\d{2})\s+"  # start
    r"(\d{1,2}:\d{2})\s+"  # end
)
_SHIFT_RANGE_RE = re.compile(r"(\d{1,2}:\d{2})-(\d{1,2}:\d{2})")


def _parse_preliminary_line(line: str, line_index: int) -> Optional[RawRow]:
    """
    Parse preliminary DOS line. Start and End are the projected shift (truth).
    Try: (1) Start/End columns as two HH:MM, or (2) shift_time range HH:MM-HH:MM.
    """
    line = line.strip()
    if not line or TRANSIT_SUPERVISOR in line:
        return None
    actual_start_str = actual_end_str = scheduled_end_str = None
    paddle = block = shift_time_str = ""
    rest_start = 0

    # Try paddle block start end (Start/End columns)
    m = _PRELIM_START_END_RE.match(line)
    if m:
        paddle, block, actual_start_str, actual_end_str = m.group(1), m.group(2), m.group(3), m.group(4)
        scheduled_end_str = actual_end_str
        shift_time_str = f"{actual_start_str}-{actual_end_str}"
        rest_start = m.end()

    # Fallback: paddle block shift_time hrs
    if not actual_start_str and not actual_end_str:
        m = _PRELIM_RE.match(line)
        if m:
            paddle, block, shift_time_str, _ = m.group(1), m.group(2), m.group(3), m.group(4)
            sm = _SHIFT_RANGE_RE.search(shift_time_str)
            if sm:
                actual_start_str, actual_end_str = sm.group(1), sm.group(2)
                scheduled_end_str = actual_end_str
                rest_start = m.end()

    if not actual_start_str or not actual_end_str:
        return None
    rest = line[rest_start:].strip()
    drivers, rest = _extract_driver_ids_and_rest(rest)
    if not drivers:
        return None
    primary_name = _strip_leading_trim(drivers[0][0])
    emp_id = drivers[0][1]
    alt_emp_id = drivers[1][1] if len(drivers) >= 2 else ""
    alt_name = _strip_leading_trim(drivers[1][0]) if len(drivers) >= 2 else ""
    notes_text, primary_condition_text = _classify_remainder(rest)
    potential_bleed = _has_potential_bleed(notes_text, emp_id)
    return RawRow(
        paddle=paddle,
        block=block,
        shift_time_str=shift_time_str,
        actual_start_str=actual_start_str,
        actual_end_str=actual_end_str,
        scheduled_end_str=scheduled_end_str,
        emp_id=emp_id,
        employee_name=primary_name.strip(),
        alternate_driver_present=len(drivers) >= 2,
        notes_text=notes_text,
        primary_condition_text=primary_condition_text,
        source_line_index=line_index,
        potential_bleed=potential_bleed,
        alternate_emp_id=alt_emp_id,
        alternate_name=alt_name,
    )


def parse_preliminary_dos_lines(lines: List[str]) -> Tuple[List[RawRow], int, Optional[str]]:
    """Parse lines as preliminary DOS. Same merging of note continuations."""
    rows = []
    stopped_at = 0
    work_date = None
    date_re = re.compile(r"(\d{1,2})/(\d{1,2})/(\d{2,4})")
    i = 0
    while i < len(lines):
        line = lines[i]
        if TRANSIT_SUPERVISOR in line:
            stopped_at = i + 1
            break
        dm = date_re.search(line)
        if dm:
            mo, day, yr = dm.group(1), dm.group(2), dm.group(3)
            if len(yr) == 2:
                yr = "20" + yr
            work_date = f"{int(mo):02d}/{int(day):02d}/{yr}"
        raw = _parse_preliminary_line(line, i)
        if raw is not None:
            j = i + 1
            extra = []
            while j < len(lines) and _is_note_continuation(lines[j]):
                extra.append(lines[j].strip())
                j += 1
            if extra:
                raw.notes_text = (raw.notes_text + " " + " ".join(extra)).strip()
            rows.append(raw)
            i = j
        else:
            i += 1
    if stopped_at == 0:
        stopped_at = len(lines)
    return rows, stopped_at, work_date


def load_preliminary_dos(path: Path) -> Tuple[List[RawRow], int, Optional[str]]:
    """Load preliminary (projected) DOS. Same logic, different column layout."""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(str(path))
    suf = path.suffix.lower()
    if suf == ".pdf":
        lines = extract_lines_from_pdf(path)
        return parse_preliminary_dos_lines(lines)
    if suf == ".csv":
        return _load_preliminary_csv(path)
    if suf in (".xlsx", ".xls"):
        return _load_preliminary_xlsx(path)
    lines = extract_lines_from_text_file(path)
    return parse_preliminary_dos_lines(lines)


def _load_preliminary_csv(path: Path) -> Tuple[List[RawRow], int, Optional[str]]:
    """Preliminary CSV: Start and End columns are the projected shift (source of truth)."""
    import csv
    rows_out = []
    work_date = None
    date_re = re.compile(r"(\d{1,2})/(\d{1,2})/(\d{2,4})")
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        reader = csv.DictReader(f)
        for i, row in enumerate(reader):
            if "transit" in str(row).lower() and "supervisor" in str(row).lower():
                return rows_out, i + 1, work_date
            prim = _col(row, "primary driver", "primary_driver", "primary")
            if not prim:
                continue
            p_name, p_id = _parse_driver_cell(prim)
            if not p_id:
                continue
            # Start/End are the projected shift — they are the truth (nothing to compare to)
            start_str = _col(row, "start")
            end_str = _col(row, "end")
            if not start_str or not end_str:
                shift = _col(row, "shift time", "shift_time", "shift")
                if shift:
                    sm = re.search(r"(\d{1,2}:\d{2})-(\d{1,2}:\d{2})", shift)
                    if sm:
                        start_str, end_str = sm.group(1), sm.group(2)
            if not start_str or not end_str:
                continue
            shift_time_str = f"{start_str}-{end_str}"
            alt = _col(row, "alternate driver", "alternate_driver", "alternate")
            alt_name, alt_id = _parse_driver_cell(alt)
            notes = _col(row, "notes")
            cond = _col(row, "primary condition", "primary_condition", "condition")
            _, primary_cond = _classify_remainder(notes + " " + cond) if (notes or cond) else ("", "")
            raw = RawRow(
                paddle=_col(row, "paddle"),
                block=_col(row, "block"),
                shift_time_str=shift_time_str,
                actual_start_str=start_str,
                actual_end_str=end_str,
                scheduled_end_str=end_str,
                emp_id=p_id,
                employee_name=p_name,
                alternate_driver_present=bool(alt_id),
                notes_text=notes,
                primary_condition_text=primary_cond,
                source_line_index=i,
                potential_bleed=_has_potential_bleed(notes, p_id),
                alternate_emp_id=alt_id,
                alternate_name=alt_name,
            )
            rows_out.append(raw)
    return rows_out, len(rows_out) + 1, work_date


def _load_preliminary_xlsx(path: Path) -> Tuple[List[RawRow], int, Optional[str]]:
    """Preliminary Excel: Start and End columns are the projected shift (source of truth)."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl required for Excel. pip install openpyxl")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], 0, None
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    header_lower = [h.lower() for h in header]
    work_date = None
    date_re = re.compile(r"(\d{1,2})/(\d{1,2})/(\d{2,4})")
    rows_out = []
    for i, row in enumerate(rows[1:], start=1):
        if not row:
            continue
        if "transit" in " ".join(str(x) for x in row).lower() and "supervisor" in " ".join(str(x) for x in row).lower():
            return rows_out, i + 1, work_date

        def col(*keys):
            for k in keys:
                for j, h in enumerate(header_lower):
                    if k in h and j < len(row) and row[j] is not None:
                        return str(row[j]).strip()
            return ""

        prim = col("primary driver", "primary_driver", "primary")
        if not prim:
            continue
        p_name, p_id = _parse_driver_cell(prim)
        if not p_id:
            continue
        # Start/End are the projected shift — they are the truth (nothing to compare to)
        start_str = col("start")
        end_str = col("end")
        if not start_str or not end_str:
            shift = col("shift time", "shift_time", "shift")
            if shift:
                sm = re.search(r"(\d{1,2}:\d{2})-(\d{1,2}:\d{2})", shift)
                if sm:
                    start_str, end_str = sm.group(1), sm.group(2)
        if not start_str or not end_str:
            continue
        shift_time_str = f"{start_str}-{end_str}"
        alt = col("alternate driver", "alternate_driver", "alternate")
        alt_name, alt_id = _parse_driver_cell(alt)
        notes = col("notes")
        cond = col("primary condition", "primary_condition", "condition")
        _, primary_cond = _classify_remainder(notes + " " + cond) if (notes or cond) else ("", "")
        raw = RawRow(
            paddle=col("paddle"),
            block=col("block"),
            shift_time_str=shift_time_str,
            actual_start_str=start_str,
            actual_end_str=end_str,
            scheduled_end_str=end_str,
            emp_id=p_id,
            employee_name=p_name,
            alternate_driver_present=bool(alt_id),
            notes_text=notes,
            primary_condition_text=primary_cond,
            source_line_index=i,
            potential_bleed=_has_potential_bleed(notes, p_id),
            alternate_emp_id=alt_id,
            alternate_name=alt_name,
        )
        rows_out.append(raw)
    return rows_out, len(rows_out) + 1, work_date


def load_dos(pdf_or_txt_path: Path) -> Tuple[List[RawRow], int, Optional[str]]:
    """Load DOS from PDF, TXT, CSV, or Excel. Returns (rows, stopped_at_1based, work_date)."""
    path = Path(pdf_or_txt_path)
    if not path.exists():
        raise FileNotFoundError(str(path))
    suf = path.suffix.lower()
    if suf == ".pdf":
        lines = extract_lines_from_pdf(path)
        return parse_dos_lines(lines)
    if suf == ".csv":
        return extract_raw_rows_from_csv(path)
    if suf in (".xlsx", ".xls"):
        return extract_raw_rows_from_xlsx(path)
    # .txt or fallback: line-oriented text
    lines = extract_lines_from_text_file(path)
    return parse_dos_lines(lines)

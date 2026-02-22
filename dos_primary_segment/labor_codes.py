"""
Load labor codes (label -> code) from Excel. Used for leave/condition codes in excluded rows.
Format: Label, Code columns (or first two columns). Sheet "codes" or active.
"""
from pathlib import Path
from typing import Dict, Optional

# Default codes when no file uploaded
DEFAULT_LABOR_CODES: Dict[str, str] = {
    "SICK": "3009",
    "FMLA SICK": "2024",
    "VACATION": "3008",
    "FMLA VACATION": "2025",
    "ADMIN LEAVE": "3010",
    "CT PAY": "3003",
}


def load_labor_codes_xlsx(path: Path, sheet_name: Optional[str] = None) -> Dict[str, str]:
    """Load label->code from Excel. Expects Label, Code columns (or first two)."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl required. pip install openpyxl")
    result = dict(DEFAULT_LABOR_CODES)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    elif "codes" in wb.sheetnames:
        ws = wb["codes"]
    elif "labor" in wb.sheetnames:
        ws = wb["labor"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return result
    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    label_idx = next((i for i, h in enumerate(header) if "label" in h or "name" in h or "desc" in h), 0)
    code_idx = next((i for i, h in enumerate(header) if "code" in h), 1)
    if code_idx == label_idx:
        code_idx = 1 if label_idx == 0 else 0
    for row in rows[1:]:
        if not row or (label_idx >= len(row) and code_idx >= len(row)):
            continue
        label = str(row[label_idx]).strip() if label_idx < len(row) and row[label_idx] else ""
        code = str(row[code_idx]).strip() if code_idx < len(row) and row[code_idx] else ""
        if label and code:
            result[label.upper()] = code
    return result


def load_labor_codes(path: Optional[Path] = None) -> Dict[str, str]:
    """Load labor codes. Returns defaults if path is None or empty."""
    if not path or not path.exists():
        return dict(DEFAULT_LABOR_CODES)
    suf = path.suffix.lower()
    if suf in (".xlsx", ".xls"):
        return load_labor_codes_xlsx(path)
    return dict(DEFAULT_LABOR_CODES)
